from openpyxl import Workbook

workbook = Workbook()

worksheet = workbook.active
worksheet.title = "New Title"
print(workbook.sheetnames)

bonus_worksheet = workbook.create_sheet("test_sheet")
print(bonus_worksheet)

cell = worksheet['A4']
worksheet['A4'] = 4
print(cell)
print(cell.value)


cell2 = worksheet.cell(row=4, column=2, value=10)
print(cell2)
print(cell2.value)


col_range = bonus_worksheet['A:B']
for i in col_range:
    for j in i:
        print(j)

row_range = bonus_worksheet[1:5]
for i in row_range:
    for j in i:
        print(j)

cell_range = bonus_worksheet['A1':'B2']
for i in cell_range:
    for j in i:
        print(j)


workbook.save('test.xlsx')